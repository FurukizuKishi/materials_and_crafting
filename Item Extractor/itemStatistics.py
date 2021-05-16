from openpyxl import load_workbook;
import item as it;
    
def get_weapon_category(title, styles, style):
    print("Checking title '%s' for style '%s'..." % (title, style), end = " ");
    if title.upper() == title:
        if style in title:
            print("Found a type 1 style.");
            return get_style_in_title(title, style);
        else:
            for s in styles:
                if s.upper() != style:
                    if s.upper() in title:
                        print("Found a type 2 style '%s'." % s);
                        return get_style_in_title(title, s.upper());
        print("Found a type 3 style.");
        return get_title_case(title), None;
    print("Could not find a style.");
    return None;

def get_style_in_title(title, style):
    t = title[title.index(style) : title.rfind(' ')];
    return (get_title_case(title[0 : title.index(style) - 1]), get_title_case(t[0:]));

def get_title_case(string):
    try:
        return string[0] + string[1:].lower();
    except IndexError:
        return string;

def obtain_weapons(sheet):
    return obtain_items(sheet, "A2", "D99", ["Melee", "Ranged"]);
def obtain_armours(sheet):
    return obtain_items(sheet, "F2", "I24", ["Armour", "Shield"]);

def obtain_items(sheet, min_cell, max_cell, styles):
    items = {};
    for style in styles:
        items[style] = {};
    title = "";
    style = styles[0];
    item = "";
    for row in sheet[min_cell : max_cell]:
        properties = [];
        new_title = None;
        i = 0;
        for cell in row:
            val = cell.value;
            if val != None:
                if i == 0:
                    try:
                        new_title, s = get_weapon_category(val, items.keys(), style.upper());
                        if s not in ["", None]:
                            style = s;
                        print("Found title '%s' in style '%s'." % (new_title, style));
                    except TypeError:
                        new_title = None;
                        item = val;
                else:
                    properties.append(val);
                i += 1;
        if new_title not in ["", None]:
            print("Created new title '%s'." % new_title);
            title = new_title;
            items[style][title] = [];
        elif len(properties) > 1:
            print("Created %s %s of style '%s'." % (title, item, style));
            items[style][title].append(it.Item(item, properties[0], properties[1], properties[2]));
        print("");
    return items;

def print_item_map(items):
    for style in items.keys():
        print(style);
        for category in items[style].keys():
            print(" " * 4 + category);
            for item in items[style][category]:
                print(" " * 8 + str(item.name));
                for component_list in item.components:
                    print(" " * 12, end = "");
                    for i in range(0, len(component_list)):
                        if i > 0:
                            print(" OR ", end = "");
                        print(str(component_list[i]), end = "");
                    print();

def obtain_items_sheet(name):
    file = load_workbook(filename = name, read_only = True);
    return file[file.sheetnames[0]];

items = obtain_items_sheet("../Items and Components.xlsx");
weapons = obtain_weapons(items);
armours = obtain_armours(items);

print_item_map(weapons);
print();
print_item_map(armours);
