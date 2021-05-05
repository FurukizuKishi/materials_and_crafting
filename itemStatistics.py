from openpyxl import load_workbook;
    
def get_weapon_category(title, styles, style):
    print("Checking title '%s' for style '%s'..." % (title, style), end = " ");
    if title.upper() == title:
        if style in title:
            print("Found a type 1 style.");
            return get_style_in_title(title, style);
        else:
            for s in styles:
                if s.upper() != style:
                    if s.upper() in title:
                        print("Found a type 2 style '%s'." % s);
                        return get_style_in_title(title, s.upper());
        print("Found a type 3 style.");
        return get_title_case(title), None;
    print("Could not find a style.");
    return None;

def get_style_in_title(title, style):
    t = title[title.index(style) : title.rfind(' ')];
    return (get_title_case(title[0 : title.index(style) - 1]), get_title_case(t[0:]));

def get_title_case(string):
    return string[0] + string[1:].lower();

def obtain_weapons(sheet):
    return obtain_items(sheet, "A2", "D99", ["Melee", "Ranged"]);
def obtain_armours(sheet):
    return obtain_items(sheet, "F2", "I24", ["Armour", "Shield"]);

def obtain_items(sheet, min_cell, max_cell, styles):
    weapons = {};
    for style in styles:
        weapons[style] = {};
    title = "";
    style = styles[0];
    weapon = "";
    for row in sheet[min_cell : max_cell]:
        properties = [];
        new_title = None;
        i = 0;
        for cell in row:
            val = cell.value;
            if val != None:
                if i == 0:
                    try:
                        new_title, s = get_weapon_category(val, weapons.keys(), style.upper());
                        if s != None:
                            style = s;
                            
                        print("Found title '%s' in style '%s'." % (new_title, style));
                    except TypeError:
                        new_title = None;
                        weapon = val;
                else:
                    properties.append(val.replace("\n", " "));
                i += 1;
        if new_title != None:
            print("Created new title '%s'." % new_title);
            title = new_title;
            weapons[style][title] = {};
        elif len(properties) > 1:
            print("Created %s %s of style '%s'." % (title, weapon, style));
            weapons[style][title][weapon] = properties;
        print("");
    return weapons;

def obtain_items_sheet(name):
    file = load_workbook(filename = name, read_only = True);
    return file[file.sheetnames[0]];

items = obtain_items_sheet("Items and Components.xlsx");
weapons = obtain_weapons(items);
armours = obtain_armours(items);

for style in weapons.keys():
    print(style);
    for category in weapons[style].keys():
        print(" " * 4 + category);
        for weapon in weapons[style][category].keys():
            print(" " * 8 + str(weapon));
            for attribute in weapons[style][category][weapon]:
                print(" " * 12 + str(attribute));
